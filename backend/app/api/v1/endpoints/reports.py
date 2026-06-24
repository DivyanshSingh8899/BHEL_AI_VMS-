"""
Reporting endpoints — PDF, Excel, and CSV report generation.
"""
import io
import csv
from datetime import datetime, timezone
from typing import Optional
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from app.core.database import get_db
from app.core.deps import get_current_active_user
from app.models.user import User
from app.models.visitor import Visitor
from app.models.entry_exit import EntryExitLog

router = APIRouter(prefix="/reports", tags=["Reports"])


async def _fetch_report_data(db: AsyncSession, date_from: str, date_to: str, department: Optional[str] = None):
    query = (
        select(Visitor, EntryExitLog)
        .join(EntryExitLog, EntryExitLog.visitor_id == Visitor.id)
        .where(
            and_(
                EntryExitLog.visit_date >= date_from,
                EntryExitLog.visit_date <= date_to,
            )
        )
        .order_by(EntryExitLog.entry_time.desc())
    )
    if department:
        query = query.where(Visitor.department_name.ilike(f"%{department}%"))

    result = await db.execute(query)
    return result.fetchall()


def _format_duration(minutes: Optional[float]) -> str:
    if not minutes:
        return "In Progress"
    h = int(minutes // 60)
    m = int(minutes % 60)
    return f"{h}h {m}m" if h > 0 else f"{m}m"


@router.get("/csv")
async def download_csv(
    date_from: str = Query(..., description="DD-MM-YYYY"),
    date_to: str = Query(..., description="DD-MM-YYYY"),
    department: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    rows = await _fetch_report_data(db, date_from, date_to, department)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Visitor ID", "Name", "Mobile", "Email", "Company",
        "Department", "Host Employee", "Purpose",
        "Visit Date", "Day", "Entry Time", "Exit Time", "Duration",
        "Entry Method", "Status"
    ])

    for row in rows:
        v, log = row.Visitor, row.EntryExitLog
        writer.writerow([
            v.visitor_id, v.name, v.mobile, v.email or "", v.company or "",
            v.department_name or "", v.host_employee_name, v.purpose,
            log.visit_date, log.visit_day,
            log.entry_time.strftime("%H:%M:%S") if log.entry_time else "",
            log.exit_time.strftime("%H:%M:%S") if log.exit_time else "",
            _format_duration(log.duration_minutes),
            log.entry_method.value if log.entry_method else "",
            v.status.value,
        ])

    output.seek(0)
    filename = f"BHEL_VMS_Report_{date_from}_to_{date_to}.csv"
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/excel")
async def download_excel(
    date_from: str = Query(..., description="DD-MM-YYYY"),
    date_to: str = Query(..., description="DD-MM-YYYY"),
    department: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    rows = await _fetch_report_data(db, date_from, date_to, department)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visitor Report"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Visitor ID", "Name", "Mobile", "Company", "Department",
        "Host Employee", "Purpose", "Visit Date", "Day",
        "Entry Time", "Exit Time", "Duration", "Status"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F0F4F8")
    for i, row in enumerate(rows):
        v, log = row.Visitor, row.EntryExitLog
        data = [
            v.visitor_id, v.name, v.mobile, v.company or "", v.department_name or "",
            v.host_employee_name, v.purpose, log.visit_date, log.visit_day,
            log.entry_time.strftime("%H:%M:%S") if log.entry_time else "",
            log.exit_time.strftime("%H:%M:%S") if log.exit_time else "",
            _format_duration(log.duration_minutes),
            v.status.value,
        ]
        ws.append(data)
        if i % 2 == 1:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    # Column widths
    col_widths = [18, 25, 15, 20, 20, 25, 30, 14, 12, 12, 12, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Title row
    ws.insert_rows(1)
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"BHEL Varanasi — Visitor Report | {date_from} to {date_to}"
    title_cell.font = Font(bold=True, size=14, color="1A3A5C")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"BHEL_VMS_Report_{date_from}_to_{date_to}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/summary")
async def report_summary(
    date_from: str = Query(...),
    date_to: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    rows = await _fetch_report_data(db, date_from, date_to)
    total = len(rows)
    completed = sum(1 for r in rows if r.EntryExitLog.exit_time is not None)
    active = total - completed
    avg_dur = sum(r.EntryExitLog.duration_minutes or 0 for r in rows if r.EntryExitLog.duration_minutes) / max(completed, 1)

    dept_count: dict = {}
    for row in rows:
        dept = row.Visitor.department_name or "Unknown"
        dept_count[dept] = dept_count.get(dept, 0) + 1

    return {
        "date_from": date_from,
        "date_to": date_to,
        "total_visits": total,
        "completed_visits": completed,
        "active_visits": active,
        "avg_duration_minutes": round(avg_dur, 1),
        "department_breakdown": dept_count,
    }
